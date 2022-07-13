from openpyxl import Workbook
from openpyxl.styles import Alignment
wb = Workbook()

sheet = wb.active
sheet.merge_cells('A1:B1')
cell = sheet.cell(row=1,column=1)
cell.value = 'Amos Wachira'
cell.alignment = Alignment(horizontal='center',vertical='center')
wb.save('merging_cells.xlsx')

